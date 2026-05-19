import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, datetime, timedelta
import os, sys, glob

# ── INSTALL DEPENDENCIES IF NEEDED ───────────────────────────────────────────
try:
    from tkcalendar import Calendar
    import tkinter as tk
except ImportError:
    import subprocess
    print("Installing required packages...")
    subprocess.check_call([sys.executable,"-m","pip","install","tkcalendar","--quiet"])
    from tkcalendar import Calendar
    import tkinter as tk

# ── CALENDAR POPUP ────────────────────────────────────────────────────────────
def pick_date():
    selected=[None]
    root=tk.Tk()
    root.title("DISREP Report — Select Date")
    root.resizable(False,False)
    root.configure(bg="#1F3864")

    tk.Label(root,text="DISREP Daily Report Generator",
             font=("Arial",13,"bold"),bg="#1F3864",fg="white").pack(pady=(15,3))
    tk.Label(root,text="Mojec International Limited",
             font=("Arial",9),bg="#1F3864",fg="#BDD7EE").pack()
    tk.Label(root,text="Select the date for your report:",
             font=("Arial",10),bg="#1F3864",fg="#BDD7EE").pack(pady=(8,5))

    cal=Calendar(root,selectmode='day',
                 year=date.today().year,month=date.today().month,day=date.today().day,
                 date_pattern='dd/mm/yyyy',font="Arial 10",
                 background="#2E75B6",foreground="white",
                 headersbackground="#1F3864",headersforeground="white",
                 selectbackground="#F4B942",selectforeground="black",
                 normalbackground="white",weekendbackground="#E2EFDA",
                 othermonthbackground="#F2F2F2")
    cal.pack(padx=20,pady=5)

    def confirm():
        selected[0]=datetime.strptime(cal.get_date(),'%d/%m/%Y').date()
        root.destroy()
    def cancel():
        root.destroy(); sys.exit()

    f=tk.Frame(root,bg="#1F3864"); f.pack(pady=(5,15))
    tk.Button(f,text="✅  Generate Report",command=confirm,
              font=("Arial",10,"bold"),bg="#F4B942",fg="black",
              padx=20,pady=8,relief="flat",cursor="hand2").pack(side="left",padx=10)
    tk.Button(f,text="✖  Cancel",command=cancel,
              font=("Arial",10),bg="#FF6B6B",fg="white",
              padx=20,pady=8,relief="flat",cursor="hand2").pack(side="left",padx=10)

    w=root.winfo_reqwidth(); h=root.winfo_reqheight()
    root.geometry(f"+{(root.winfo_screenwidth()-w)//2}+{(root.winfo_screenheight()-h)//2}")
    root.update_idletasks(); root.mainloop()
    return selected[0]

print("="*55)
print("  DISREP DAILY REPORT GENERATOR")
print("  Prepared by: Haruna Idris | Mojec International")
print("="*55)
print()

REPORT_DATE=pick_date()
if not REPORT_DATE:
    sys.exit()

print(f"Report date: {REPORT_DATE.strftime('%d-%b-%Y')}")
print()
print("⏳ Processing data...")
print()

# ── FIND FDM FILES ────────────────────────────────────────────────────────────
script_dir=os.path.dirname(os.path.abspath(__file__))

def find_file(keywords):
    all_xlsx=[f for f in glob.glob(os.path.join(script_dir,"*.xlsx"))
              if not os.path.basename(f).startswith('DISREP_Daily')]
    for f in all_xlsx:
        name=os.path.basename(f).lower()
        if any(k.lower() in name for k in keywords):
            return f
    return None

install_file=find_file(['install plan','install_plan','installation'])
survey_file =find_file(['survey plan','survey_plan','survey'])

if not install_file:
    print("❌ ERROR: Install Plan file not found in folder!")
    input("\nPress Enter to close..."); sys.exit()
if not survey_file:
    print("❌ ERROR: Survey Plan file not found in folder!")
    input("\nPress Enter to close..."); sys.exit()

print(f"✅ {os.path.basename(install_file)}")
print(f"✅ {os.path.basename(survey_file)}")
print()

# ── LOAD DATA ─────────────────────────────────────────────────────────────────
df_inst=pd.read_excel(install_file)
df_surv=pd.read_excel(survey_file)

if 'TeamName' in df_inst.columns: df_inst.rename(columns={'TeamName':'Team Name'},inplace=True)
if 'TeamName' in df_surv.columns: df_surv.rename(columns={'TeamName':'Team Name'},inplace=True)

df_inst['Upload Time']=pd.to_datetime(df_inst['Upload Time'],errors='coerce')
df_inst['Upload Date']=df_inst['Upload Time'].dt.date
df_inst['Upload Month']=df_inst['Upload Time'].dt.month
df_inst['Upload Year']=df_inst['Upload Time'].dt.year
for col in ['Work Order Status','Meter Type','Team Name','Feeder Name']:
    df_inst[col]=df_inst[col].astype(str).str.strip()

df_surv['Update Time']=pd.to_datetime(df_surv['Update Time'],errors='coerce')
df_surv['Update Date']=df_surv['Update Time'].dt.date
df_surv['Update Month']=df_surv['Update Time'].dt.month
df_surv['Update Year']=df_surv['Update Time'].dt.year
for col in ['Work Order Status','Team Name','Feeder Name']:
    df_surv[col]=df_surv[col].astype(str).str.strip()

# ── REGION DETECTION ──────────────────────────────────────────────────────────
JIGAWA_KW =['birnin kudu','sani abacha','takur','government house','school of nursing']
KATSINA_KW=['jibia','hassan usman','low cost','dandagoro']

def get_region(f):
    fl=str(f).lower()
    if any(k in fl for k in JIGAWA_KW):  return 'Jigawa'
    if any(k in fl for k in KATSINA_KW): return 'Katsina'
    return 'Kano'

df_inst['Region']=df_inst['Feeder Name'].apply(get_region)
df_surv['Region']=df_surv['Feeder Name'].apply(get_region)

DONE=['Confirmed','Rejected','To Be Confirmed']
def sc(mask): return int(mask.sum())

# ── INSTALLATION METRICS ──────────────────────────────────────────────────────
df_id=df_inst[df_inst['Work Order Status'].isin(DONE)]
df_it=df_id[df_id['Upload Date']==REPORT_DATE]
df_ia=df_inst[df_inst['Work Order Status']=='Assigned']

im={}
im['done_so_far']        =len(df_id)
im['confirmed']          =sc(df_inst['Work Order Status']=='Confirmed')
im['to_be_confirm']      =sc(df_inst['Work Order Status']=='To Be Confirmed')
im['rejected']           =sc(df_inst['Work Order Status']=='Rejected')
im['nov']=sc((df_id['Upload Month']==11)&(df_id['Upload Year']==2025))
im['dec']=sc((df_id['Upload Month']==12)&(df_id['Upload Year']==2025))
im['jan']=sc((df_id['Upload Month']==1) &(df_id['Upload Year']==2026))
im['feb']=sc((df_id['Upload Month']==2) &(df_id['Upload Year']==2026))
im['mar']=sc((df_id['Upload Month']==3) &(df_id['Upload Year']==2026))
im['apr']=sc((df_id['Upload Month']==4) &(df_id['Upload Year']==2026))
im['may']=sc((df_id['Upload Month']==5) &(df_id['Upload Year']==2026))
im['kano_total']         =sc(df_id['Region']=='Kano')
im['katsina_total']      =sc(df_id['Region']=='Katsina')
im['jigawa_total']       =sc(df_id['Region']=='Jigawa')
im['today_total']        =len(df_it)
im['today_kano']         =sc(df_it['Region']=='Kano')
im['today_katsina']      =sc(df_it['Region']=='Katsina')
im['today_jigawa']       =sc(df_it['Region']=='Jigawa')
im['single_phase']       =sc(df_it['Meter Type']=='S12U16')
im['three_phase']        =sc(df_it['Meter Type']=='S34U18')
im['installable_kano']   =sc(df_ia['Region']=='Kano')
im['installable_katsina']=sc(df_ia['Region']=='Katsina')
im['installable_jigawa'] =sc(df_ia['Region']=='Jigawa')
im['total_installable']  =len(df_ia)
im['total_kyc_upload'] = 80061

# Total meters allocated
im['total_allocated']    =len(df_inst)

# Project completion %
# Contract volume
CONTRACT_VOLUME = 128800

# Project completion %
im['completion_pct'] = round(
    (im['done_so_far'] / CONTRACT_VOLUME * 100), 1
) if CONTRACT_VOLUME > 0 else 0

# SBC today only
sbc_today={}
for t in sorted(df_inst['Team Name'].dropna().unique()):
    cnt=sc(df_it['Team Name']==t)
    if cnt>0: sbc_today[t]=cnt

# Feeder today
feeder_today=df_it.groupby('Feeder Name').size().sort_values(ascending=False).to_dict()

# Rejection reasons today
rej_today={}
if 'Reject Reason' in df_inst.columns:
    df_rej=df_inst[(df_inst['Work Order Status']=='Rejected')&(df_inst['Upload Date']==REPORT_DATE)]
    if len(df_rej)>0:
        rej_today=df_rej['Reject Reason'].astype(str).str.strip().value_counts().to_dict()

# Rejection reasons cumulative
rej_all={}
if 'Reject Reason' in df_inst.columns:
    df_rej_all=df_inst[df_inst['Work Order Status']=='Rejected']
    if len(df_rej_all)>0:
        rej_all=df_rej_all['Reject Reason'].astype(str).str.strip().value_counts().head(8).to_dict()

# Top 5 SBC cumulative
top5_sbc=df_id.groupby('Team Name').size().sort_values(ascending=False).head(5).to_dict()

# ── SURVEY METRICS ────────────────────────────────────────────────────────────
df_sd=df_surv[df_surv['Work Order Status'].isin(DONE)]
df_st=df_sd[df_sd['Update Date']==REPORT_DATE]
df_sa=df_surv[df_surv['Work Order Status']=='Assigned']

sm={}
sm['done_so_far']        =len(df_sd)
sm['confirmed']          =sc(df_surv['Work Order Status']=='Confirmed')
sm['to_be_confirm']      =sc(df_surv['Work Order Status']=='To Be Confirmed')
sm['rejected']           =sc(df_surv['Work Order Status']=='Rejected')
sm['nov']=sc((df_sd['Update Month']==11)&(df_sd['Update Year']==2025))
sm['dec']=sc((df_sd['Update Month']==12)&(df_sd['Update Year']==2025))
sm['jan']=sc((df_sd['Update Month']==1) &(df_sd['Update Year']==2026))
sm['feb']=sc((df_sd['Update Month']==2) &(df_sd['Update Year']==2026))
sm['mar']=sc((df_sd['Update Month']==3) &(df_sd['Update Year']==2026))
sm['apr']=sc((df_sd['Update Month']==4) &(df_sd['Update Year']==2026))
sm['may']=sc((df_sd['Update Month']==5) &(df_sd['Update Year']==2026))
sm['kano_total']         =sc(df_sd['Region']=='Kano')
sm['katsina_total']      =sc(df_sd['Region']=='Katsina')
sm['jigawa_total']       =sc(df_sd['Region']=='Jigawa')
sm['today_total']        =len(df_st)
sm['today_kano']         =sc(df_st['Region']=='Kano')
sm['today_katsina']      =sc(df_st['Region']=='Katsina')
sm['today_jigawa']       =sc(df_st['Region']=='Jigawa')
sm['assignable_kano']    =sc(df_sa['Region']=='Kano')
sm['assignable_katsina'] =sc(df_sa['Region']=='Katsina')
sm['assignable_jigawa']  =sc(df_sa['Region']=='Jigawa')

# Survey vs Installation gap
sm['survey_inst_gap']=sm['done_so_far']-im['done_so_far']

# ── STYLES ────────────────────────────────────────────────────────────────────
DB="1F3864"; MB="2E75B6"; LB="BDD7EE"
DG="375623"; LG="E2EFDA"
OR="F4B942"; YL="FFF2CC"; WH="FFFFFF"; GR="F2F2F2"
RD="FF0000"; LR="FFCCCC"

thin=Side(style='thin'); thick=Side(style='medium')
fb=Border(left=thin,right=thin,top=thin,bottom=thin)
tb=Border(left=thick,right=thick,top=thick,bottom=thick)

def s(ws,cell,val=None,bg=None,fg="000000",bold=False,sz=10,
      align="center",wrap=False,nf=None,italic=False,bdr='thin'):
    c=ws[cell] if isinstance(cell,str) else cell
    if val is not None: c.value=val
    c.font=Font(bold=bold,color=fg,size=sz,name="Arial",italic=italic)
    if bg: c.fill=PatternFill("solid",start_color=bg)
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if nf: c.number_format=nf
    c.border=fb
    return c

def h(ws,cell,val,bg=DB,fg=WH,sz=10):
    s(ws,cell,val,bg=bg,fg=fg,bold=True,sz=sz)

def sec_title(ws,merge,cell,val,bg=MB, fg=WH):
    ws.merge_cells(merge)
    s(ws,cell,val,bg=bg,fg=fg,bold=True,sz=10,align="center")

DATE_STR =REPORT_DATE.strftime('%d-%b-%Y').upper()
DATE_LONG=REPORT_DATE.strftime('%d-%B, %Y').upper()

# ══════════════════════════════════════════════════════════════════════════════
wb=Workbook()
ws=wb.active
ws.title=REPORT_DATE.strftime('%d-%m-%Y')
ws.sheet_view.showGridLines=False

# Column widths
for col,w in zip(['A','B','C','D','E','F','G','H','I','J'],
                 [32,13,3,6,24,13,3,6,24,13]):
    ws.column_dimensions[col].width=w

# ── MAIN TITLE ────────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
s(ws,"A1",f"{DATE_LONG} REPORT — MOJEC INTERNATIONAL LIMITED | DISREP METERING PROJECT",
  bg=DB,fg=WH,bold=True,sz=12,align="center")
ws.row_dimensions[1].height=28

# ── PROJECT CONFIGURATION ─────────────────────────────────────────
CONTRACT_VOLUME = 128800

# Remaining meters
im['remaining_meters'] = CONTRACT_VOLUME - im['done_so_far']

# Project completion %
im['completion_pct'] = round(
    (im['done_so_far'] / CONTRACT_VOLUME * 100), 1
) if CONTRACT_VOLUME > 0 else 0


# ── PROJECT COMPLETION BANNER ─────────────────────────────────────
ws.merge_cells("A2:F2")

completion_text = (
    f"PROJECT COMPLETION: {im['completion_pct']}%  |  "
    f"Contract Volume: {CONTRACT_VOLUME:,}  |  "
    f"Installed: {im['done_so_far']:,}  |  "
    f"Remaining: {im['remaining_meters']:,}  |  "
)

s(
    ws,
    "A2",
    completion_text,
    bg=OR,
    fg=DB,
    bold=True,
    sz=10,
    align="center"
)

ws.row_dimensions[2].height = 20

R=4  # current row tracker

# ══════════════════════════════════════════════════════════════════════════════
# ROW 1: SURVEY (left) | SURVEY REGION (right)
# ══════════════════════════════════════════════════════════════════════════════
sec_title(ws,f"A{R}:B{R}",f"A{R}","SURVEY SUMMARY",bg=MB)
sec_title(ws,f"D{R}:F{R}",f"D{R}","SURVEY BY REGION",bg=MB)
ws.row_dimensions[R].height=18; R+=1

survey_left=[
    ("Total Surveyed so far",    sm['done_so_far']),
    ("Survey Confirm on FDM",    sm['confirmed']),
    ("Survey yet To be Confirm", sm['to_be_confirm']),
    ("Survey Rejected on FDM",   sm['rejected']),
    ("Survey done in January",   sm['jan']),
    ("Survey done in February",  sm['feb']),
    ("Survey done in March",     sm['mar']),
    ("Survey done in April",     sm['apr']),
    ("Survey done in May",       sm['may']),
    ("Survey done today",        sm['today_total']),
]
survey_region=[
    ("Total Survey Done In Kano",        sm['kano_total']),
    ("Total Survey Done In Katsina",     sm['katsina_total']),
    ("Total Survey Done In Jigawa",      sm['jigawa_total']),
    ("Yet To Survey in Kano",     sm['assignable_kano']),
    ("Yet To Survey in Katsina",  sm['assignable_katsina']),
    ("Yet To Survey in Jigawa",   sm['assignable_jigawa']),
    ("Survey done Today in Kano",             sm['today_kano']),
    ("Survey done Today in Katsina",          sm['today_katsina']),
    ("Survey done Today in Jigawa",           sm['today_jigawa']),
    ("Survey vs Install Gap",    sm['survey_inst_gap']),
]
for i,(( lm,lv),(rm,rv)) in enumerate(zip(survey_left,survey_region)):
    row=R+i; bg=GR if i%2==0 else WH
    s(ws,f"A{row}",lm,bg=bg,align="left",sz=10)
    s(ws,f"B{row}",lv,bg=LB,bold=True,nf="#,##0")
    s(ws,f"D{row}",rm,bg=bg,align="left",sz=10)
    gap_bg=LG if rv>=0 else LR
    s(ws,f"E{row}",rm,bg=bg,align="left",sz=10)
    s(ws,f"F{row}",rv,bg=gap_bg if rm=="Survey vs Install Gap" else LB,
      bold=True,nf="#,##0")
    ws.row_dimensions[row].height=17

R+=len(survey_left)+1

# ══════════════════════════════════════════════════════════════════════════════
# ROW 2: INSTALLATION (left) | INSTALL REGION (right)
# ══════════════════════════════════════════════════════════════════════════════
sec_title(ws,f"A{R}:B{R}",f"A{R}","INSTALLATION SUMMARY",bg=DG,fg=WH)
sec_title(ws,f"D{R}:F{R}",f"D{R}","INSTALLATION BY REGION",bg=DG,fg=WH)
ws.row_dimensions[R].height=18; R+=1

install_left=[
    ("Total KYC Uploaded on FDM",           im['total_kyc_upload']),
    ("Installation done so far",            im['done_so_far']),
    ("Confirmed on FDM",                    im['confirmed']),
    ("Installation yet To be Confirm",      im['to_be_confirm']),
    ("Installation Rejected on FDM",        im['rejected']),
    ("Installation done in January",        im['jan']),
    ("Installation done in February",       im['feb']),
    ("Installation done in March",          im['mar']),
    ("Installation done in April",          im['apr']),
    ("Installation done in May",            im['may']),
    ("Installation done today",             im['today_total']),
]
install_region=[
    ("Total Meters Installed in Kano",       im['kano_total']),
    ("Total Meters Installed in Katsina",    im['katsina_total']),
    ("Total Meters Installed in Jigawa",     im['jigawa_total']),
    ("Total Installable for Kano",        im['installable_kano']),
    ("Total Installable for Katsina",     im['installable_katsina']),
    ("Total Installable for Jigawa",      im['installable_jigawa']),
    ("Installation done today in Kano",            im['today_kano']),
    ("Installation done today in Katsina",         im['today_katsina']),
    ("Installation done today in Jigawa",          im['today_jigawa']),
    ("Single Phase Installed today",      im['single_phase']),
    ("Three Phase Installed today",       im['three_phase']),
]

max_rows=max(len(install_left),len(install_region))
for i in range(max_rows):
    row=R+i; bg=GR if i%2==0 else WH
    if i<len(install_left):
        lm,lv=install_left[i]
        s(ws,f"A{row}",lm,bg=bg,align="left",sz=10)
        s(ws,f"B{row}",lv,bg=LG,bold=True,nf="#,##0")
    if i<len(install_region):
        rm,rv=install_region[i]
        s(ws,f"D{row}",rm,bg=bg,align="left",sz=10)
        val_bg=LG if isinstance(rv,(int,float)) else YL
        s(ws,f"F{row}",rv,bg=val_bg,bold=True,
          nf="#,##0" if isinstance(rv,(int,float)) else None)
    ws.row_dimensions[row].height=17

R+=max_rows+1

# ══════════════════════════════════════════════════════════════════════════════
# ROW 3: SBC TODAY (left) | FEEDER TODAY (right)
# ══════════════════════════════════════════════════════════════════════════════
sec_title(ws,f"A{R}:B{R}",f"A{R}",f"SBC INSTALLATION — {DATE_STR}",bg=OR,fg=DB)
sec_title(ws,f"D{R}:F{R}",f"D{R}",f"FEEDER INSTALLATION — {DATE_STR}",bg=OR,fg=DB)
ws.row_dimensions[R].height=18; R+=1

# Header row
h(ws,f"A{R}","SBC / TEAM NAME",bg=OR,fg=DB)
h(ws,f"B{R}","COUNT",         bg=OR,fg=DB)
h(ws,f"D{R}","FEEDER NAME",   bg=OR,fg=DB)
h(ws,f"F{R}","COUNT",         bg=OR,fg=DB)
ws.row_dimensions[R].height=17; R+=1

sbc_items=sorted(sbc_today.items(),key=lambda x:x[1],reverse=True)
feeder_items=sorted(feeder_today.items(),key=lambda x:x[1],reverse=True)
max_r=max(len(sbc_items),len(feeder_items),1)

for i in range(max_r):
    row=R+i; bg=GR if i%2==0 else WH
    if i<len(sbc_items):
        t,c=sbc_items[i]
        s(ws,f"A{row}",t,bg=bg,align="left",sz=10)
        s(ws,f"B{row}",c,bg=LG,bold=True,nf="#,##0")
    if i<len(feeder_items):
        f2,c2=feeder_items[i]
        s(ws,f"D{row}",f2,bg=bg,align="left",sz=10)
        s(ws,f"F{row}",c2,bg=LB,bold=True,nf="#,##0")
    ws.row_dimensions[row].height=16

# Totals
tot_row=R+max_r
if sbc_items:
    s(ws,f"A{tot_row}","TOTAL",bg=OR,fg=DB,bold=True)
    s(ws,f"B{tot_row}",sum(sbc_today.values()),bg=OR,bold=True,nf="#,##0")
if feeder_items:
    s(ws,f"D{tot_row}","TOTAL",bg=OR,fg=DB,bold=True)
    s(ws,f"F{tot_row}",sum(feeder_today.values()),bg=OR,bold=True,nf="#,##0")
ws.row_dimensions[tot_row].height=17

if not sbc_items:
    ws.merge_cells(f"A{R}:B{R}")
    s(ws,f"A{R}",f"No installations on {DATE_STR}",bg=YL,bold=True,align="center")
if not feeder_items:
    ws.merge_cells(f"D{R}:F{R}")
    s(ws,f"D{R}",f"No feeder data on {DATE_STR}",bg=YL,bold=True,align="center")

R=tot_row+2

# ── FOOTER ────────────────────────────────────────────────────────────────────
ws.merge_cells(f"A{R}:F{R}")
s(ws,f"A{R}",
  f"Prepared by: Haruna Idris | Data Analyst | Mojec International Limited | DISREP Metering Project | {DATE_STR}",
  bg=DB,fg=WH,bold=False,sz=9,align="center",italic=True)
ws.row_dimensions[R].height=16

# ── SAVE & OPEN ───────────────────────────────────────────────────────────────
output=os.path.join(script_dir,
    f"DISREP_Daily_Report_{REPORT_DATE.strftime('%d_%b_%Y')}.xlsx")
wb.save(output)

print(f"✅ SURVEY: Done={sm['done_so_far']:,} | Today={sm['today_total']:,}")
print(f"✅ INSTALL: Done={im['done_so_far']:,} | Today={im['today_total']:,}")
print(f"✅ Project: {im['completion_pct']}% complete")
print(f"✅ SBC active today: {len(sbc_items)}")
print(f"✅ Feeders active today: {len(feeder_items)}")
print()
print(f"✅ Saved: {os.path.basename(output)}")
print()

os.startfile(output)
input("Press Enter to close...")
